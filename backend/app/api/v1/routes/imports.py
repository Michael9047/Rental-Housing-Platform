import io
import os

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user, get_db_session, require_landlord
from app.models.user import User
from app.services.audit_service import AuditService
from app.services.import_service import ImportService

router = APIRouter()



ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}
MAX_UPLOAD_SIZE = 10 * 1024 * 1024  # 10 MB


def _validate_upload(file: UploadFile) -> str:
    if not file.filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No filename provided",
        )
    _, ext = os.path.splitext(file.filename)
    ext = ext.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported file type '{ext}'. Allowed: {ALLOWED_EXTENSIONS}",
        )
    return ext


def _map_ext_to_source_type(ext: str) -> str:
    if ext == ".csv":
        return "csv"
    return "excel"


@router.post("/upload")
async def upload_import(
    file: UploadFile,
    session: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_landlord),
    institute_id: int | None = Query(default=None, description="公寓ID，批量导入的所有房源归属此公寓"),
    mode: str | None = Query(default="flexible", description="导入模式: flexible=跳过错误行, strict=有错全部回滚"),
) -> dict:
    ext = _validate_upload(file)
    content = await file.read()
    if len(content) == 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="文件无内容，请使用下载的模板填写数据后重新上传",
        )
    if len(content) > MAX_UPLOAD_SIZE:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"File too large. Maximum size: {MAX_UPLOAD_SIZE // (1024*1024)} MB",
        )

    source_type = _map_ext_to_source_type(ext)
    import_service = ImportService(session)

    import_task = await import_service.create_import_task(
        admin_id=current_user.id,
        source_name=file.filename,
        source_type=source_type,
    )

    import_task = await import_service.parse_and_import(
        import_task=import_task,
        file_content=content,
        landlord_id=current_user.id,
        institute_id=institute_id,
        mode=mode or "flexible",
    )

    await AuditService(session).create_log(
        user_id=current_user.id,
        action="data_import",
        resource_type="import",
        resource_id=import_task.id,
        details={
            "source_name": file.filename,
            "source_type": source_type,
            "total": import_task.total_records,
            "success": import_task.success_records,
            "failed": import_task.failed_records,
        },
    )

    import json

    error_log = import_task.error_log
    if error_log:
        try:
            error_log = json.loads(error_log)
        except json.JSONDecodeError:
            pass

    return {
        "id": import_task.id,
        "source_name": import_task.source_name,
        "source_type": import_task.source_type.value,
        "status": import_task.status.value,
        "total_records": import_task.total_records,
        "success_records": import_task.success_records,
        "failed_records": import_task.failed_records,
        "error_log": error_log,
        "created_at": import_task.created_at.isoformat(),
    }


@router.get("/tasks")
async def list_tasks(
    session: AsyncSession = Depends(get_db_session),
    _: User = Depends(require_landlord),
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=200),
    status: str | None = Query(default=None),
) -> list[dict]:
    tasks = await ImportService(session).list_tasks(
        skip=skip, limit=limit, status=status,
    )
    return [
        {
            "id": t.id,
            "admin_id": t.admin_id,
            "source_name": t.source_name,
            "source_type": t.source_type.value,
            "status": t.status.value,
            "total_records": t.total_records,
            "success_records": t.success_records,
            "failed_records": t.failed_records,
            "created_at": t.created_at.isoformat(),
        }
        for t in tasks
    ]


@router.get("/tasks/{task_id}")
async def get_task_detail(
    task_id: int,
    session: AsyncSession = Depends(get_db_session),
    _: User = Depends(require_landlord),
) -> dict:
    task = await ImportService(session).get_import_task(task_id)
    if not task:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import task not found")

    import json

    error_log = task.error_log
    if error_log:
        try:
            error_log = json.loads(error_log)
        except json.JSONDecodeError:
            pass

    return {
        "id": task.id,
        "admin_id": task.admin_id,
        "source_name": task.source_name,
        "source_type": task.source_type.value,
        "status": task.status.value,
        "total_records": task.total_records,
        "success_records": task.success_records,
        "failed_records": task.failed_records,
        "error_log": error_log,
        "created_at": task.created_at.isoformat(),
        "updated_at": task.updated_at.isoformat(),
    }


@router.post("/tasks/{task_id}/retry")
async def retry_failed_records(
    task_id: int,
    session: AsyncSession = Depends(get_db_session),
    current_user: User = Depends(require_landlord),
) -> dict:
    task = await ImportService(session).get_import_task(task_id)
    if not task:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import task not found")

    if task.status.value not in {"completed", "failed"}:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot retry task with status '{task.status.value}'",
        )

    task = await ImportService(session).retry_failed(
        import_task=task, landlord_id=current_user.id,
    )

    import json

    error_log = task.error_log
    if error_log:
        try:
            error_log = json.loads(error_log)
        except json.JSONDecodeError:
            pass

    return {
        "id": task.id,
        "source_name": task.source_name,
        "status": task.status.value,
        "total_records": task.total_records,
        "success_records": task.success_records,
        "failed_records": task.failed_records,
        "error_log": error_log,
        "updated_at": task.updated_at.isoformat(),
    }


# ── 模板下载 ────────────────────────────────────────────

@router.get("/template")
async def download_template() -> StreamingResponse:
    """
    下载标准 Excel 导入模板

    包含两个 Sheet:
      Sheet 1 "房源导入模板" — 带表头、示例数据、下拉验证
      Sheet 2 "填写说明" — 字段含义、枚举值对照、常见错误提示
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed. Run: pip install openpyxl",
        )

    wb = openpyxl.Workbook()

    # ── 样式定义 ────────────────────────────────────────
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    required_font = Font(name="微软雅黑", size=11, bold=True, color="CC0000")
    optional_font = Font(name="微软雅黑", size=11, color="333333")
    example_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    # ── Sheet 1: 房源导入模板 ────────────────────────────
    ws1 = wb.active
    ws1.title = "房源导入模板"

    # 列定义: (字段名, 中文名, 是否必填, 示例值, 列宽, 说明)
    columns = [
        ("title",           "房源标题*",     True,  "工业园区翰林缘精装单间",        22, "必填，200字以内，建议格式：区域+小区+户型亮点"),
        ("address",         "详细地址*",     True,  "苏州工业园区仁爱路199号翰林缘",  28, "必填，含路名+小区名+门牌号"),
        ("district",        "所在区域*",     True,  "工业园区",                      14, "必填，苏州区域名"),
        ("price_monthly",   "月租金(元)*",   True,  "2500",                          14, "必填，数字，单位元/月"),
        ("area_sqm",        "面积(㎡)",      False, "80",                            12, "选填，数字"),
        ("bedrooms",        "卧室数",        False, "2",                             10, "选填，整数，如2室1厅填2"),
        ("bathrooms",       "卫生间数",      False, "1",                             10, "选填，整数"),
        ("property_type",   "房源类型",      False, "apartment",                     14, "选填: apartment/house/studio/shared"),
        ("description",     "房源描述",      False, "精装修，家电齐全，适合留学生",   30, "选填"),
        ("deposit_amount",  "押金(元)",      False, "2500",                          14, "选填，数字"),
        ("service_fee_rate","服务费比例",    False, "0.10",                          12, "选填，如10%填0.10"),
        ("building_name",   "楼栋名称",      False, "翰林缘公寓",                    18, "选填，用于自动关联已有楼栋"),
        ("room_number",     "房号",          False, "1201",                          10, "选填"),
        ("floor",           "楼层",          False, "12",                            10, "选填，整数"),
        ("latitude",        "纬度",          False, "31.315",                        14, "选填"),
        ("longitude",       "经度",          False, "120.715",                       14, "选填"),
    ]

    # 表头行
    for col_idx, (field, label, required, example, width, desc) in enumerate(columns, 1):
        cell = ws1.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    ws1.row_dimensions[1].height = 36

    # 示例数据行
    for col_idx, (field, label, required, example, width, desc) in enumerate(columns, 1):
        cell = ws1.cell(row=2, column=col_idx, value=example)
        cell.font = Font(name="微软雅黑", size=10, color="888888")
        cell.fill = example_fill
        cell.border = thin_border

    # 预留 3 行空行
    for row in range(3, 6):
        for col_idx in range(1, len(columns) + 1):
            cell = ws1.cell(row=row, column=col_idx, value="")
            cell.border = thin_border
            cell.font = Font(name="微软雅黑", size=10)

    # 下拉验证 — 房源类型
    type_dv = DataValidation(
        type="list",
        formula1='"apartment,house,studio,shared"',
        allow_blank=True,
    )
    type_dv.error = "请选择: apartment, house, studio, shared"
    type_dv.errorTitle = "无效房源类型"
    type_col_letter = get_column_letter(8)  # 房源类型在第8列
    ws1.add_data_validation(type_dv)
    type_dv.add(f"{type_col_letter}2:{type_col_letter}1000")

    # 下拉验证 — 区域
    district_dv = DataValidation(
        type="list",
        formula1='"工业园区,姑苏区,高新区,吴中区,相城区,吴江区,昆山市,太仓市,常熟市,张家港市"',
        allow_blank=True,
    )
    district_dv.error = "请选择有效的苏州区域"
    district_dv.errorTitle = "无效区域"
    district_col_letter = get_column_letter(3)
    ws1.add_data_validation(district_dv)
    district_dv.add(f"{district_col_letter}2:{district_col_letter}1000")

    # 冻结首行
    ws1.freeze_panes = "A2"

    # ── Sheet 2: 填写说明 ────────────────────────────────
    ws2 = wb.create_sheet("填写说明")

    title_font = Font(name="微软雅黑", size=14, bold=True, color="FF6B35")
    section_font = Font(name="微软雅黑", size=11, bold=True)
    normal_font = Font(name="微软雅黑", size=10)

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 50
    ws2.column_dimensions["C"].width = 40

    row = 1
    ws2.cell(row=row, column=1, value="📋 房源批量导入模板 — 填写说明").font = title_font
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    ws2.cell(row=row, column=1, value="一、必填字段（标 * 号）").font = section_font
    row += 1
    required_fields = [
        ("房源标题", '格式建议：区域+小区名+户型亮点，如"工业园区翰林缘精装单间"'),
        ("详细地址", "含路名+小区名+门牌号，用于地理编码和地图展示"),
        ("所在区域", "必须从下拉列表中选择，用于区域筛选和租金统计"),
        ("月租金(元)", '纯数字，不含"元"字，如 2500。系统会自动检测异常值'),
    ]
    for label, desc in required_fields:
        ws2.cell(row=row, column=1, value=label).font = Font(name="微软雅黑", size=10, bold=True, color="CC0000")
        ws2.cell(row=row, column=2, value=desc).font = normal_font
        row += 1

    row += 1
    ws2.cell(row=row, column=1, value="二、房源类型枚举").font = section_font
    row += 1
    type_enum = [
        ("apartment", "普通公寓/住宅"),
        ("house", "别墅/独栋"),
        ("studio", "单间/单身公寓/开间"),
        ("shared", "合租/合住"),
    ]
    for val, desc in type_enum:
        ws2.cell(row=row, column=1, value=val).font = Font(name="Consolas", size=10, bold=True)
        ws2.cell(row=row, column=2, value=desc).font = normal_font
        row += 1

    row += 1
    ws2.cell(row=row, column=1, value="三、注意事项").font = section_font
    row += 1
    tips = [
        "1. 文件格式：支持 .csv / .xlsx / .xls，最大 10MB",
        '2. 列名自动匹配：支持中英文列名，如"标题"→title、"月租金"→price_monthly',
        "3. 编码支持：自动识别 UTF-8 / GBK / ISO-8859-1，海外导出的表格不会乱码",
        "4. 异常检测：系统会自动标记异常租金和面积值，请在导入结果中核实",
        "5. 重复检测：相同标题+地址的房源会自动跳过",
        "6. 楼栋关联：如填写楼栋名称，系统会自动匹配已有楼栋",
        "7. 批量导入后会自动生成 AI 周边分析和语义搜索索引，请稍等片刻",
        "8. 有疑问请联系平台管理员",
    ]
    for tip in tips:
        ws2.cell(row=row, column=1, value=tip).font = normal_font
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1

    # ── 输出 ────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=property_import_template.xlsx",
        },
    )


@router.get("/tasks/{task_id}/errors/download")
async def download_error_table(
    task_id: int,
    session: AsyncSession = Depends(get_db_session),
    _: User = Depends(require_landlord),
) -> StreamingResponse:
    """
    下载仅含错误行的 Excel 表格
    保留原始内容 + 新增 error_type(错误类型) + suggest(整改建议) 列
    """
    import json as _json
    task = await ImportService(session).get_import_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")

    error_log = task.error_log
    if not error_log:
        raise HTTPException(status_code=400, detail="该任务无错误记录")

    try:
        errors = _json.loads(error_log) if isinstance(error_log, str) else error_log
    except _json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="错误记录解析失败")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "错误行"

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    err_fill = PatternFill(start_color="FFF2F2", end_color="FFF2F2", fill_type="solid")

    ws.append(["原始行号", "错误类型", "错误原因", "整改建议"])
    for ci in range(1, 5):
        cell = ws.cell(1, ci)
        cell.font = header_font; cell.fill = header_fill

    suggest_map = {
        "duplicate": "请确认是否与已有房源重复，如是同一房源请勿重复导入",
        "missing_field": "请在原始文件中补充缺失的必填字段后重新上传",
        "format_error": "请按模板要求修正数据格式（数字/枚举/日期）",
    }

    for e in errors:
        if e.get("row", 0) == 0:
            continue
        row = e.get("row", 0)
        etype = e.get("type", "unknown")
        ws.append([
            row,
            {"duplicate":"重复","missing_field":"缺字段","format_error":"格式错"}.get(etype, "其他"),
            e.get("error", ""),
            suggest_map.get(etype, "请检查该行数据是否填写完整且格式正确"),
        ])

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 45

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=errors_task{task_id}.xlsx"},
    )
